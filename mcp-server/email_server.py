"""Email MCP Server based on FastMCP
提供邮件发送功能的 MCP 服务
"""
import json
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from typing import Optional, List
from pathlib import Path
from datetime import datetime, timedelta
from pydantic import BaseModel, Field

from mcp.server.fastmcp import FastMCP

mcp = FastMCP("email-server")


@mcp.tool()
async def create_weekly_report(json_data_file: str, key_desc_file: str, output_file: Optional[str] = None) -> dict:
    """
    创建周报 Excel
    
    Args:
        json_data_file: JSON 数据文件路径
        key_desc_file: Key 描述文件路径
        output_file: 输出 Excel 文件路径（可选）
    
    Returns:
        生成结果
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        with open(json_data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        with open(key_desc_file, 'r', encoding='utf-8') as f:
            key_descriptions = json.load(f)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "周报"
        
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        column_order = ['weekly_time', 'requirement_name', 'requirement_id', 'is_cosmic_metric',
                       'product_manager', 'developer', 'total_task_man_days', 'used_task_man_days',
                       'requirement_description', 'execution_role', 'standard_hours', 'actual_hours',
                       'priority', 'requirement_status', 'other_notes', 'planned_completion_date',
                       'actual_completion_date', 'work_content_progress', 'launch_date', 'is_external_support']
        
        headers = [key_descriptions.get(key, key) for key in column_order]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(column_order, 1):
                value = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
        
        if output_file:
            wb.save(output_file)
            return {"success": True, "output_file": output_file}
        
        return {"success": True, "message": "周报创建成功"}
    
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
async def send_email(subject: str, content: str, to_emails: List[str], 
                    attachment_path: Optional[str] = None) -> dict:
    """
    发送邮件
    
    Args:
        subject: 邮件主题
        content: 邮件内容
        to_emails: 收件人邮箱列表
        attachment_path: 附件路径（可选）
    
    Returns:
        发送结果
    """
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.exmail.qq.com')
        smtp_port = int(os.getenv('SMTP_PORT', '465'))
        sender_email = os.getenv('SENDER_EMAIL')
        sender_password = os.getenv('SENDER_PASSWORD')
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = subject
        
        msg.attach(MIMEText(content, 'html', 'utf-8'))
        
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
                msg.attach(part)
        
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_emails, msg.as_string())
        server.quit()
        
        return {"success": True, "message": "邮件发送成功"}
    
    except Exception as e:
        return {"success": False, "error": str(e)}
